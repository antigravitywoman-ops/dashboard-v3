#!/usr/bin/env python3
"""
EXCEL PORTER SKILL — Real Implementation
Reads .md markdown table files and generates a formatted .xlsx via openpyxl.
Default source: memory/sheets/*.md (agent-driven path).
Override with --source=<path> for report-generator.js path (e.g. reports/2026-W11/sheets/).
Requires: pip install openpyxl
"""

import sys
import os
import re
import json
from datetime import datetime, timezone
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[excel-porter] ERROR: Missing openpyxl. Run: pip install openpyxl")
    sys.exit(1)


# ── Format Rules ──────────────────────────────────────────────────────────────

SHEET_CONFIG = {
    "01-executive-summary":   {"tab": "1F3864", "header": "1F3864", "widths": {"A": 30, "B": 60}},
    "02-gap-analysis":        {"tab": "E06C00", "header": "7F3F00", "widths": {"A": 40, "B": 20, "C": 20, "D": 30}},
    "03-competitor-analysis": {"tab": "C00000", "header": "4C0000", "widths": {"A": 35, "B": 30, "C": 12, "D": 12, "E": 25}},
    "04-twelve-week-plan":    {"tab": "203864", "header": "1F3864", "widths": {"A": 8, "B": 40, "C": 15, "D": 15, "E": 15, "F": 25}},
    "05-keyword-research":    {"tab": "375623", "header": "1E4620", "widths": {"A": 40, "B": 14, "C": 12, "D": 12, "E": 14, "F": 18}},
    "06-location-pages":      {"tab": "7030A0", "header": "4B0082", "widths": {"A": 30, "B": 50, "C": 20, "D": 30}},
    "07-citations-backlinks":   {"tab": "00B0F0", "header": "00457C", "widths": {"A": 35, "B": 25, "C": 15, "D": 15}},
    "08-youtube-strategy":    {"tab": "FF0000", "header": "8B0000", "widths": {"A": 40, "B": 30, "C": 15, "D": 10}},
    "09-reddit-quora":        {"tab": "FF4500", "header": "7A2000", "widths": {"A": 30, "B": 40, "C": 15, "D": 10}},
    "10-review-strategy":     {"tab": "FFC000", "header": "7F6000", "widths": {"A": 30, "B": 40, "C": 15, "D": 20}},
    "11-schema-markup":       {"tab": "70AD47", "header": "375623", "widths": {"A": 30, "B": 20, "C": 15, "D": 50}},
    "12-weekly-tasks":        {"tab": "ED7D31", "header": "843C0C", "widths": {"A": 6, "B": 35, "C": 20, "D": 12, "E": 15, "F": 25}},
    "13-kpis-metrics":        {"tab": "7030A0", "header": "2C0066", "widths": {"A": 30, "B": 20, "C": 20, "D": 20, "E": 15}},
}

DISPLAY_NAMES = {
    "01-executive-summary":   "Executive Summary",
    "02-gap-analysis":        "Gap Analysis",
    "03-competitor-analysis": "Competitor Analysis",
    "04-twelve-week-plan":    "12-Week Plan",
    "05-keyword-research":    "Keyword Research",
    "06-location-pages":      "Location Pages",
    "07-citations-backlinks":   "Citations & Backlinks",
    "08-youtube-strategy":    "YouTube Strategy",
    "09-reddit-quora":        "Reddit & Quora",
    "10-review-strategy":     "Review Strategy",
    "11-schema-markup":       "Schema Markup",
    "12-weekly-tasks":        "Weekly Tasks",
    "13-kpis-metrics":        "KPIs & Metrics",
}

# Fills
WHITE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def parse_md_table(md_content: str) -> tuple[list, list]:
    """
    Parse the first markdown table found in the content.
    Returns (headers: list[str], rows: list[dict]).
    """
    lines = md_content.splitlines()
    table_lines = []
    in_table = False

    for line in lines:
        stripped = line.strip()
        if re.match(r'^\|.+\|$', stripped):
            in_table = True
            table_lines.append(stripped)
        elif in_table and stripped == '':
            break
        elif in_table and not re.match(r'^\|', stripped):
            break

    if len(table_lines) < 2:
        return [], []

    # Parse headers from first row
    headers = [h.strip() for h in table_lines[0].strip('|').split('|')]
    # Skip separator line (row 1 which is |---|---|)
    rows = []
    for line in table_lines[2:]:
        cells = [c.strip() for c in line.strip('|').split('|')]
        row = dict(zip(headers, cells[:len(headers)]))
        rows.append(row)

    return headers, rows


def apply_conditional_formatting(ws, sheet_key: str, headers: list, data_start: int, data_end: int):
    """Apply sheet-specific conditional formatting rules."""
    header_lower = [h.lower() for h in headers]

    status_col = None
    for i, h in enumerate(header_lower):
        if 'status' in h:
            status_col = get_column_letter(i + 1)
            break

    position_col = None
    for i, h in enumerate(header_lower):
        if 'position' in h or 'pos' in h:
            position_col = get_column_letter(i + 1)
            break

    priority_col = None
    for i, h in enumerate(header_lower):
        if 'priority' in h:
            priority_col = get_column_letter(i + 1)
            break

    # Status-based row coloring (applied manually per row)
    if status_col:
        for row_num in range(data_start, data_end + 1):
            cell = ws[f"{status_col}{row_num}"]
            val = str(cell.value or "").strip().lower()
            if val in ("done", "submitted", "implemented"):
                cell.fill = GREEN_FILL
            elif val in ("in progress", "weak"):
                cell.fill = YELLOW_FILL
            elif val in ("missing", "blocked", "critical"):
                cell.fill = RED_FILL

    # Position-based coloring for keyword sheets
    if position_col and sheet_key == "05-keyword-research":
        for row_num in range(data_start, data_end + 1):
            cell = ws[f"{position_col}{row_num}"]
            try:
                pos = float(cell.value or 0)
                if 0 < pos <= 3:
                    cell.fill = GREEN_FILL
                elif 4 <= pos <= 10:
                    cell.fill = YELLOW_FILL
                elif pos > 20:
                    cell.fill = RED_FILL
            except (ValueError, TypeError):
                pass

    # Priority-based coloring
    if priority_col:
        for row_num in range(data_start, data_end + 1):
            cell = ws[f"{priority_col}{row_num}"]
            val = str(cell.value or "").strip().upper()
            if val in ("H", "HIGH"):
                cell.fill = RED_FILL
            elif val in ("M", "MEDIUM", "MED"):
                cell.fill = YELLOW_FILL


def write_sheet(ws, sheet_key: str, headers: list, rows: list):
    """Write headers and data rows to a worksheet with full formatting."""
    config = SHEET_CONFIG.get(sheet_key, {"header": "1F3864", "widths": {}})
    header_fill = make_fill(config["header"])
    col_widths = config.get("widths", {})

    # Set tab color
    ws.sheet_properties.tabColor = config.get("tab", "1F3864")

    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = WHITE_FONT
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            # Try numeric conversion
            try:
                num = float(val.replace(",", "").replace("%", "")) if val else val
                if "%" in str(val):
                    cell = ws.cell(row=row_idx, column=col_idx, value=num / 100)
                    cell.number_format = "0.0%"
                elif "." in str(val):
                    cell = ws.cell(row=row_idx, column=col_idx, value=num)
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value=int(num) if num == int(num) else num)
                    header_l = header.lower()
                    if any(k in header_l for k in ["volume", "clicks", "impressions", "sessions"]):
                        cell.number_format = "#,##0"
            except (ValueError, AttributeError):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Zebra striping
            if row_idx % 2 == 0:
                cell.fill = ZEBRA_FILL

    data_end = len(rows) + 1

    # Apply conditional formatting
    apply_conditional_formatting(ws, sheet_key, headers, 2, data_end)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[col_letter]
        else:
            ws.column_dimensions[col_letter].width = 18


def port_to_excel(company_slug: str, single_sheet: str = None, source_dir: str = None):
    print(f"[excel-porter] Starting Excel generation for: {company_slug}")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Default to memory/sheets/; override via --source for report-generator path
    sheets_dir = source_dir or os.path.join(script_dir, '..', '..', '..', 'companies', company_slug, 'memory', 'sheets')
    report_dir = os.path.join(script_dir, '..', '..', '..', 'companies', company_slug, 'reports')
    os.makedirs(report_dir, exist_ok=True)

    if not os.path.isdir(sheets_dir):
        print(f"[excel-porter] ERROR: sheets directory not found: {sheets_dir}")
        sys.exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    # Determine which sheets to process
    keys_to_process = list(SHEET_CONFIG.keys())
    if single_sheet:
        single_clean = single_sheet.replace(".md", "")
        keys_to_process = [k for k in keys_to_process if k == single_clean]
        if not keys_to_process:
            print(f"[excel-porter] ERROR: Unknown sheet key: {single_sheet}")
            sys.exit(1)

    sheets_built = 0
    sheets_skipped = 0

    for sheet_key in keys_to_process:
        md_file = os.path.join(sheets_dir, f"{sheet_key}.md")
        display_name = DISPLAY_NAMES.get(sheet_key, sheet_key)

        if not os.path.exists(md_file):
            print(f"[excel-porter] SKIP (missing): {sheet_key}.md")
            sheets_skipped += 1
            continue

        with open(md_file, 'r', encoding='utf-8') as f:
            content = f.read()

        headers, rows = parse_md_table(content)
        if not headers:
            print(f"[excel-porter] SKIP (no table found): {sheet_key}.md")
            sheets_skipped += 1
            continue

        ws = wb.create_sheet(title=display_name)
        write_sheet(ws, sheet_key, headers, rows)
        print(f"[excel-porter]   ✓ {display_name} — {len(rows)} rows, {len(headers)} columns")
        sheets_built += 1

    if sheets_built == 0:
        print("[excel-porter] ERROR: No sheets were built. Check that memory/sheets/*.md files exist with tables.")
        sys.exit(1)

    date_str = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    file_name = f"SEO_Strategy_{company_slug}_{date_str}.xlsx"
    out_file = os.path.join(report_dir, file_name)
    wb.save(out_file)

    print(f"\n[excel-porter] SUCCESS: {sheets_built} sheets generated ({sheets_skipped} skipped)")
    print(f"[excel-porter] Output: {out_file}")

    # Append to episodic log
    episodic_path = os.path.join(script_dir, '..', '..', '..', 'companies', company_slug, 'memory', 'episodic.md')
    if os.path.exists(episodic_path):
        with open(episodic_path, 'a', encoding='utf-8') as ep:
            ep.write(f"\n- [{date_str}] Excel-porter generated: `{file_name}` ({sheets_built} sheets)\n")

    return {
        "success": True,
        "company": company_slug,
        "sheets_generated": sheets_built,
        "sheets_skipped": sheets_skipped,
        "report_path": out_file,
        "generated_at": datetime.now(timezone.utc).isoformat()
    }


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        slug = sys.argv[1]
        single = None
        source = None
        for arg in sys.argv[2:]:
            if arg.startswith("--sheet="):
                single = arg.split("=")[1]
            elif arg.startswith("--source="):
                source = arg.split("=", 1)[1]
        result = port_to_excel(slug, single, source)
        print(json.dumps(result, indent=2))
    else:
        print("Usage: python3 excel-porter.py <company_slug> [--sheet=<sheet-key>] [--source=<path>]")
        print("  --source=<path>  Override default memory/sheets/ path (e.g. reports/2026-W11/sheets/)")
        print("\nSheet keys:")
        for k, v in DISPLAY_NAMES.items():
            print(f"  {k}  →  {v}")
