#!/usr/bin/env python3
"""
EXCEL PARSER SKILL — Real Implementation
Parses a client .xlsx file using pandas (cell values) and openpyxl (colors + formulas).
Requires: pip install pandas openpyxl python-dotenv
"""

import sys
import json
import os
from datetime import datetime, timezone

try:
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[excel-parser] ERROR: Missing dependencies. Run: pip install pandas openpyxl")
    sys.exit(1)


def hex_from_color(color):
    """Extract hex string from openpyxl Color object."""
    if color is None:
        return None
    if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
        return '#' + color.rgb[-6:]  # strip alpha prefix
    return None


def extract_formatting(wb, sheet_name: str) -> dict:
    """Extract cell colors and formulas from openpyxl workbook."""
    ws = wb[sheet_name]
    highlighted_rows = set()
    column_colors = {}
    formulas = {}

    for row in ws.iter_rows():
        for cell in row:
            col_letter = get_column_letter(cell.column)
            # Extract fill color
            fill = cell.fill
            if fill and fill.fgColor:
                hex_color = hex_from_color(fill.fgColor)
                if hex_color and hex_color.lower() not in ('#ffffff', '#000000'):
                    highlighted_rows.add(cell.row)
                    if col_letter not in column_colors:
                        column_colors[col_letter] = hex_color
            # Extract formula
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell.coordinate] = cell.value

    return {
        "highlightedRows": sorted([r - 1 for r in highlighted_rows if r > 1]),  # 0-indexed data rows
        "columnColors": column_colors,
        "formulas": formulas
    }


def parse_excel_deliverable(file_path: str, company_slug: str) -> dict:
    print(f"[excel-parser] Parsing: {file_path}")
    print(f"[excel-parser] Company: {company_slug}")

    if not os.path.exists(file_path):
        print(f"[excel-parser] ERROR: File not found: {file_path}")
        sys.exit(1)

    # Pass 1: pandas for cell values
    print(f"[excel-parser] Pass 1: Reading cell values via pandas...")
    all_sheets_dfs = pd.read_excel(file_path, sheet_name=None, header=0, dtype=str)

    # Pass 2: openpyxl for colors + formulas (data_only=False to read formula strings)
    print(f"[excel-parser] Pass 2: Reading colors and formulas via openpyxl...")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)  # for computed values

    sheets_output = {}
    for sheet_name, df in all_sheets_dfs.items():
        print(f"[excel-parser]   → Sheet: {sheet_name} ({len(df)} rows)")
        df = df.where(pd.notna(df), None)  # replace NaN with None
        headers = list(df.columns)
        rows = df.to_dict(orient='records')

        # Get formatting if sheet exists in workbook
        formatting = {}
        if sheet_name in wb.sheetnames:
            formatting = extract_formatting(wb, sheet_name)

        sheets_output[sheet_name] = {
            "headers": headers,
            "rows": rows,
            "rowCount": len(rows),
            "formatting": formatting
        }

    intelligence_payload = {
        "success": True,
        "company": company_slug,
        "sourceFile": os.path.basename(file_path),
        "sheetsParsed": len(sheets_output),
        "sheets": sheets_output,
        "parsedAt": datetime.now(timezone.utc).isoformat()
    }

    # Save to company memory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    mem_dir = os.path.join(script_dir, '..', '..', '..', 'companies', company_slug, 'memory')
    os.makedirs(mem_dir, exist_ok=True)

    out_file = os.path.join(mem_dir, 'intake-state.json')
    with open(out_file, 'w', encoding='utf-8') as f:
        json.dump(intelligence_payload, f, indent=4, ensure_ascii=False, default=str)

    print(f"[excel-parser] SUCCESS: {len(sheets_output)} sheets parsed → {out_file}")
    return intelligence_payload


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        result = parse_excel_deliverable(sys.argv[1], sys.argv[2])
        # Print summary (not full data to avoid overwhelming stdout)
        summary = {
            "success": result["success"],
            "sheetsParsed": result["sheetsParsed"],
            "sheets": {k: {"rowCount": v["rowCount"], "headers": v["headers"][:5]}
                       for k, v in result["sheets"].items()}
        }
        print(json.dumps(summary, indent=2))
    else:
        print("Usage: python3 excel-parser.py <path_to_xlsx> <company_slug>")
